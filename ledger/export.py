"""Take the ledger out of the app: a spreadsheet, or a statement to print.

Both formats are built from the same rows so the two downloads can never
disagree with each other or with the screen.

Amounts are written as plain decimal numbers, never as text with a symbol.
A spreadsheet whose money column is text is a spreadsheet you cannot sum,
which defeats the point of exporting it.
"""

from __future__ import annotations

import io
from datetime import date

from ledger.compute import by_person, ledger_breakdown, totals  # noqa: F401
from ledger.models import Entry
from ledger.money import Currency

#: The PDF core fonts are latin-1, so "₹" cannot be drawn without shipping a
#: font file. The currency code says the same thing and always renders.
PDF_SYMBOLS = {Currency.INR: "INR", Currency.USD: "USD"}

COLUMN_HEADINGS = [
    "Date", "Person", "Ledger", "Direction", "Amount", "Currency", "Note", "Attachment",
]


def _rows(entries: list[Entry]) -> list[list]:
    """One list per entry, in COLUMN_HEADINGS order, amounts as real numbers."""
    return [
        [
            entry.date,
            entry.person,
            entry.ledger,
            entry.direction.value,
            entry.amount_minor / 100,
            entry.currency.value,
            entry.note,
            entry.attachment,
        ]
        for entry in sorted(entries, key=lambda e: (e.date, e.person))
    ]


def to_excel(entries: list[Entry]) -> bytes:
    """A workbook: every entry, plus a summary sheet per currency.

    Returns the file's bytes so the caller can hand them straight to a
    download button without touching the filesystem.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    sheet = book.active
    sheet.title = "Entries"

    header_fill = PatternFill("solid", fgColor="1B2A41")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(COLUMN_HEADINGS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in _rows(entries):
        sheet.append(row)

    for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row[0].number_format = "yyyy-mm-dd"
        row[4].number_format = "#,##0.00"
    sheet.freeze_panes = "A2"

    widths = [12, 22, 22, 11, 14, 10, 34, 34]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # One summary tab per currency. They are never combined: netting rupees
    # against dollars would need an exchange rate this app refuses to invent.
    for currency in Currency:
        subset = [e for e in entries if e.currency is currency]
        if not subset:
            continue
        tab = book.create_sheet(f"Summary {currency.value}")
        tab.append(["Person", "Given", "Received", "Net owed", "Last activity", "Ledgers"])
        for cell in tab[1]:
            cell.fill = header_fill
            cell.font = header_font
        for summary in by_person(subset, currency):
            tab.append([
                summary.person,
                summary.given_minor / 100,
                summary.received_minor / 100,
                summary.net_minor / 100,
                summary.last_activity,
                summary.ledgers,
            ])
        overall = totals(subset, currency)
        tab.append([])
        tab.append([
            "TOTAL",
            overall.given_minor / 100,
            overall.received_minor / 100,
            overall.net_minor / 100,
        ])
        for row in tab.iter_rows(min_row=2):
            for cell in row[1:4]:
                cell.number_format = "#,##0.00"
            row[4].number_format = "yyyy-mm-dd"
        for index, width in enumerate([22, 14, 14, 14, 14, 10], start=1):
            tab.column_dimensions[get_column_letter(index)].width = width
        tab[tab.max_row][0].font = Font(bold=True)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _money(minor: int, currency: Currency) -> str:
    whole, frac = divmod(abs(minor), 100)
    sign = "-" if minor < 0 else ""
    return f"{PDF_SYMBOLS[currency]} {sign}{whole:,}.{frac:02d}"


def to_pdf(entries: list[Entry], *, today: date | None = None) -> bytes:
    """A statement you could hand to someone: totals, then every entry."""
    from fpdf import FPDF

    today = today or date.today()
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 10, "Personal Ledger", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(110, 110, 110)
    pdf.cell(0, 6, f"As of {today:%d %B %Y}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    if not entries:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, "No entries yet.", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf.output())

    for currency in Currency:
        subset = [e for e in entries if e.currency is currency]
        if not subset:
            continue

        overall = totals(subset, currency)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 9, currency.label, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(
            0, 6,
            f"Given {_money(overall.given_minor, currency)}    "
            f"Received {_money(overall.received_minor, currency)}    "
            f"Net outstanding {_money(overall.net_minor, currency)}",
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(2)

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(27, 42, 65)
        pdf.set_text_color(255, 255, 255)
        widths = [24, 46, 46, 24, 34, 96]
        for heading, width in zip(
            ["Date", "Person", "Ledger", "Direction", "Amount", "Note"], widths
        ):
            pdf.cell(width, 8, heading, border=0, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        pdf.set_font("Helvetica", "", 9)
        for index, entry in enumerate(sorted(subset, key=lambda e: (e.date, e.person))):
            if index % 2:
                pdf.set_fill_color(244, 244, 241)
                fill = True
            else:
                fill = False
            note = entry.note if len(entry.note) <= 58 else entry.note[:55] + "..."
            cells = [
                f"{entry.date:%d %b %Y}",
                entry.person[:26],
                entry.ledger[:26],
                entry.direction.value,
                _money(entry.amount_minor, currency),
                note,
            ]
            for text, width in zip(cells, widths):
                pdf.cell(width, 7, text, border=0, fill=fill)
            pdf.ln()

        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, f"Still owed: {_money(overall.net_minor, currency)}",
                 new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

    return bytes(pdf.output())


def demo() -> None:
    """Self-check: both files build, and the money survives the trip."""
    from ledger.models import Direction

    entries = [
        Entry(date=date(2026, 1, 5), person="Nanna", ledger="Home",
              direction=Direction.given, amount_minor=500_000, currency=Currency.INR,
              note="UPI"),
        Entry(date=date(2026, 2, 1), person="Nanna", ledger="Home",
              direction=Direction.received, amount_minor=150_000, currency=Currency.INR),
        Entry(date=date(2026, 3, 1), person="Ravi", ledger="Books",
              direction=Direction.given, amount_minor=4_250, currency=Currency.USD),
    ]

    book = to_excel(entries)
    assert book[:2] == b"PK", "not a zip, so not an xlsx"
    assert len(book) > 4000, len(book)

    pdf = to_pdf(entries, today=date(2026, 8, 22))
    assert pdf[:5] == b"%PDF-", pdf[:20]
    assert len(pdf) > 1000, len(pdf)

    # Empty ledgers must still produce openable files, not crash the page.
    assert to_pdf([])[:5] == b"%PDF-"
    assert to_excel([])[:2] == b"PK"

    # Amounts leave as numbers, not strings, or the column will not sum.
    rows = _rows(entries)
    assert rows[0][4] == 5000.0, rows[0][4]
    assert all(isinstance(row[4], float) for row in rows)

    # And the PDF text renders without a rupee glyph it cannot draw.
    assert _money(500_000, Currency.INR) == "INR 5,000.00"
    assert _money(-150_000, Currency.INR) == "INR -1,500.00"
    assert _money(4_250, Currency.USD) == "USD 42.50"

    print("ledger.export: all checks passed")


if __name__ == "__main__":
    demo()


# ------------------------------------------------------------------- sharing
# WhatsApp and mailto links carry text, not files. Neither wa.me nor mailto can
# attach anything, so what gets shared is a written summary and the person is
# told plainly that the spreadsheet has to be sent the usual way.

from urllib.parse import quote  # noqa: E402 — kept beside the code that uses it


def summary_text(entries: list[Entry], *, today: date | None = None) -> str:
    """A short, readable statement of the ledger, for pasting into a message."""
    today = today or date.today()
    if not entries:
        return f"Personal Ledger — {today:%d %b %Y}\nNothing recorded yet."

    lines = [f"Personal Ledger — as of {today:%d %b %Y}", ""]
    for currency in Currency:
        subset = [e for e in entries if e.currency is currency]
        if not subset:
            continue
        overall = totals(subset, currency)
        lines.append(f"{currency.label}")
        lines.append(f"  Given     {_money(overall.given_minor, currency)}")
        lines.append(f"  Received  {_money(overall.received_minor, currency)}")
        lines.append(f"  Still owed {_money(overall.net_minor, currency)}")
        lines.append("")
        for summary in by_person(subset, currency):
            if summary.net_minor == 0:
                continue
            owes = "owes me" if summary.net_minor > 0 else "I owe"
            lines.append(
                f"  {summary.person}: {owes} "
                f"{_money(abs(summary.net_minor), currency)}"
                f"  (last {summary.last_activity:%d %b %Y})"
            )
        lines.append("")
    return "\n".join(lines).rstrip()


def whatsapp_link(text: str) -> str:
    """A wa.me link that opens WhatsApp with the message pre-written."""
    return f"https://wa.me/?text={quote(text)}"


def email_link(text: str, subject: str = "Personal Ledger") -> str:
    """A mailto: link with the summary as the body."""
    return f"mailto:?subject={quote(subject)}&body={quote(text)}"
