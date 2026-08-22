"""Get readable content out of whatever file someone uploads.

A statement arrives as a PDF, a spreadsheet, a CSV, or a photo. Three of those
are text and should go to the text model, which reads them far better than a
vision model reads a picture of them. Only an actual image needs the vision
endpoint, and then it has to be small enough to inline.

Nothing here talks to a model. This turns bytes into either text or a
right-sized image, and the assistant decides what to do with it.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

#: The vision endpoint inlines the image as a data URI and rejects large ones.
#: Rather than refuse the upload, shrink until it fits — a statement is legible
#: at far less than phone-camera resolution.
VISION_BUDGET = 170 * 1024

#: Enough of a document to hold a statement page or two. Past this the model
#: starts losing the middle anyway, and the user is better served by a crop.
MAX_TEXT_CHARS = 24_000

TEXT_SUFFIXES = (".pdf", ".xlsx", ".xlsm", ".csv", ".txt", ".tsv")
IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp")

ACCEPTED = [s.lstrip(".") for s in (*IMAGE_SUFFIXES, *TEXT_SUFFIXES)]


class UnreadableDocument(RuntimeError):
    """The file cannot be turned into anything a model can read."""


@dataclass(frozen=True)
class Readable:
    """Either text pulled from a document, or an image small enough to send."""

    kind: str            # "text" or "image"
    text: str = ""
    data: bytes = b""
    mimetype: str = ""
    #: What happened on the way, worth showing the user.
    note: str = ""


def suffix_of(filename: str) -> str:
    name = str(filename or "").lower()
    return name[name.rfind(".") :] if "." in name else ""


def shrink_image(data: bytes, budget: int = VISION_BUDGET) -> tuple[bytes, str, str]:
    """Bring an image under `budget` bytes, returning (data, mimetype, note).

    Re-encodes as JPEG and steps the resolution down until it fits. A bank
    statement stays perfectly legible long after the file has stopped being
    four megabytes, so refusing the upload was the wrong call.
    """
    from PIL import Image

    original = len(data)
    if original <= budget:
        return data, "image/png", ""

    try:
        image = Image.open(io.BytesIO(data))
        image.load()
    except Exception as exc:  # noqa: BLE001 — any decode failure is the same to us
        raise UnreadableDocument(f"That image could not be opened ({exc}).") from exc

    if image.mode not in ("RGB", "L"):
        image = image.convert("RGB")

    width, height = image.size
    for scale in (1.0, 0.8, 0.65, 0.5, 0.4, 0.3, 0.22, 0.15):
        candidate = image
        if scale < 1.0:
            candidate = image.resize(
                (max(1, int(width * scale)), max(1, int(height * scale))),
                Image.LANCZOS,
            )
        for quality in (85, 70, 55, 40):
            buffer = io.BytesIO()
            candidate.save(buffer, format="JPEG", quality=quality, optimize=True)
            out = buffer.getvalue()
            if len(out) <= budget:
                note = (
                    f"Resized from {original // 1024} KB to {len(out) // 1024} KB "
                    f"({candidate.size[0]}×{candidate.size[1]}) so the model could read it."
                )
                return out, "image/jpeg", note

    raise UnreadableDocument(
        "That image is too dense to shrink for the vision model. "
        "Crop it to just the transactions and try again."
    )


def _pdf_text(data: bytes) -> str:
    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        raise UnreadableDocument(f"That PDF could not be opened ({exc}).") from exc

    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:  # noqa: BLE001 — one bad page must not lose the rest
            pages.append("")
    text = "\n".join(pages).strip()
    if not text:
        raise UnreadableDocument(
            "That PDF has no text in it — it is probably a scan. Screenshot the "
            "page and upload that instead; the image reader can handle it."
        )
    return text


def _sheet_text(data: bytes) -> str:
    from openpyxl import load_workbook

    try:
        book = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise UnreadableDocument(f"That spreadsheet could not be opened ({exc}).") from exc

    lines: list[str] = []
    for sheet in book.worksheets:
        # The sheet name is only worth writing if the sheet has something in
        # it; otherwise a blank workbook reads as content and never trips the
        # "this is empty" check below.
        rows = [
            " | ".join("" if c is None else str(c) for c in row)
            for row in sheet.iter_rows(values_only=True)
            if any(str(c).strip() for c in row if c is not None)
        ]
        if rows:
            lines.append(f"# sheet: {sheet.title}")
            lines.extend(rows)
    text = "\n".join(lines).strip()
    if not text:
        raise UnreadableDocument("That spreadsheet is empty.")
    return text


def _delimited_text(data: bytes) -> str:
    raw = data.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(raw[:4000])
    except csv.Error:
        return raw
    rows = list(csv.reader(io.StringIO(raw), dialect))
    return "\n".join(" | ".join(r) for r in rows if any(c.strip() for c in r))


def read(filename: str, data: bytes, mimetype: str = "") -> Readable:
    """Turn one uploaded file into something a model can be given."""
    if not data:
        raise UnreadableDocument(f"{filename} is empty.")

    suffix = suffix_of(filename)
    kind = (mimetype or "").lower()

    if suffix in IMAGE_SUFFIXES or kind.startswith("image/"):
        shrunk, out_type, note = shrink_image(data)
        return Readable(kind="image", data=shrunk, mimetype=out_type, note=note)

    if suffix == ".pdf" or "pdf" in kind:
        text = _pdf_text(data)
    elif suffix in (".xlsx", ".xlsm") or "spreadsheet" in kind or "excel" in kind:
        text = _sheet_text(data)
    elif suffix in (".csv", ".tsv") or "csv" in kind:
        text = _delimited_text(data)
    elif suffix == ".txt" or kind.startswith("text/"):
        text = data.decode("utf-8", errors="replace")
    else:
        raise UnreadableDocument(
            f"I cannot read {suffix or 'that file type'}. "
            f"Try one of: {', '.join(ACCEPTED)}."
        )

    note = ""
    if len(text) > MAX_TEXT_CHARS:
        text = text[:MAX_TEXT_CHARS]
        note = (
            f"Only the first {MAX_TEXT_CHARS:,} characters were read — the file "
            "is long. Split it if something near the end is missing."
        )
    return Readable(kind="text", text=text.strip(), note=note)


def demo() -> None:
    """Self-check: the branching, and that shrinking actually shrinks."""
    import zipfile

    # CSV keeps its shape.
    csv_bytes = b"date,person,amount\n2026-01-01,Ravi,2500\n2026-02-01,Amma,900\n"
    got = read("statement.csv", csv_bytes, "text/csv")
    assert got.kind == "text"
    assert "Ravi" in got.text and "2500" in got.text

    # Plain text passes through.
    assert read("note.txt", b"gave 500 to amma", "text/plain").text == "gave 500 to amma"

    # An unknown type says so instead of guessing.
    try:
        read("thing.dmg", b"xx", "application/octet-stream")
    except UnreadableDocument as exc:
        assert "cannot read" in str(exc)
    else:
        raise AssertionError("unknown types must be refused")

    # Empty input is refused before anything tries to parse it.
    for name in ("a.pdf", "a.csv", "a.png"):
        try:
            read(name, b"")
        except UnreadableDocument:
            pass
        else:
            raise AssertionError(f"{name}: empty should raise")

    # A junk PDF fails with a readable message, not a stack trace.
    try:
        read("broken.pdf", b"not a pdf at all", "application/pdf")
    except UnreadableDocument:
        pass
    else:
        raise AssertionError("a broken PDF should raise")

    # A big image is shrunk rather than refused — the actual bug being fixed.
    from PIL import Image

    big = Image.new("RGB", (3000, 2200))
    for x in range(0, 3000, 7):          # noise, so it cannot compress to nothing
        for y in range(0, 2200, 11):
            big.putpixel((x, y), (x % 256, y % 256, (x + y) % 256))
    buffer = io.BytesIO()
    big.save(buffer, format="PNG")
    raw = buffer.getvalue()
    assert len(raw) > VISION_BUDGET, len(raw)

    out = read("photo.png", raw, "image/png")
    assert out.kind == "image"
    assert len(out.data) <= VISION_BUDGET, len(out.data)
    assert out.mimetype == "image/jpeg"
    assert "Resized" in out.note

    # A small image is left exactly as it was.
    small = io.BytesIO()
    Image.new("RGB", (40, 40), (10, 20, 30)).save(small, format="PNG")
    untouched = read("small.png", small.getvalue(), "image/png")
    assert untouched.data == small.getvalue()
    assert untouched.note == ""

    # Long text is truncated, and says so.
    long_text = ("x" * (MAX_TEXT_CHARS + 500)).encode()
    trimmed = read("long.txt", long_text, "text/plain")
    assert len(trimmed.text) == MAX_TEXT_CHARS
    assert "first" in trimmed.note

    print("ledger.docs: all checks passed")


if __name__ == "__main__":
    demo()
