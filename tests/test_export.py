"""The downloads. What matters is that the money survives the trip."""

from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook

from ledger.export import _money, _rows, to_excel, to_pdf
from ledger.models import Direction, Entry
from ledger.money import Currency


def entry(minor: int, direction=Direction.given, currency=Currency.INR, **kw) -> Entry:
    fields = dict(
        date=date(2026, 1, 5), person="Amma", ledger="Home",
        direction=direction, amount_minor=minor, currency=currency, note="",
    )
    fields.update(kw)
    return Entry(**fields)


def sheet_of(entries, name="Entries"):
    return load_workbook(io.BytesIO(to_excel(entries)))[name]


def test_workbook_is_a_real_xlsx():
    assert to_excel([entry(500_00)])[:2] == b"PK"


def test_pdf_is_a_real_pdf():
    assert to_pdf([entry(500_00)])[:5] == b"%PDF-"


def test_amounts_are_numbers_so_the_column_sums():
    """A money column stored as text is a column you cannot add up, which
    defeats the point of exporting a ledger to a spreadsheet at all."""
    cell = sheet_of([entry(120_050)]).cell(row=2, column=5)
    assert isinstance(cell.value, (int, float))
    assert float(cell.value) == pytest.approx(1200.50)
    assert cell.number_format == "#,##0.00"


def test_dates_are_dates_not_strings():
    cell = sheet_of([entry(100, date=date(2026, 3, 9))]).cell(row=2, column=1)
    assert hasattr(cell.value, "year")
    assert (cell.value.year, cell.value.month, cell.value.day) == (2026, 3, 9)


def test_every_column_is_exported():
    from ledger.export import COLUMN_HEADINGS

    assert [c.value for c in sheet_of([entry(100)])[1]] == COLUMN_HEADINGS


def test_attachment_link_survives():
    row = _rows([entry(100, attachment="https://drive.google.com/file/d/x/view")])[0]
    assert row[7] == "https://drive.google.com/file/d/x/view"


def test_each_currency_gets_its_own_summary_tab():
    book = load_workbook(io.BytesIO(to_excel([
        entry(100, currency=Currency.INR),
        entry(200, currency=Currency.USD),
    ])))
    assert "Summary INR" in book.sheetnames
    assert "Summary USD" in book.sheetnames


def test_a_currency_with_no_entries_gets_no_tab():
    book = load_workbook(io.BytesIO(to_excel([entry(100, currency=Currency.INR)])))
    assert "Summary USD" not in book.sheetnames


def test_summary_nets_given_against_received():
    tab = sheet_of(
        [entry(1000), entry(400, direction=Direction.received)], name="Summary INR"
    )
    rows = [r for r in tab.iter_rows(values_only=True) if r[0] == "Amma"]
    _, given, received, net, *_ = rows[0]
    assert (given, received, net) == (10.0, 4.0, 6.0)


def test_empty_ledger_still_produces_openable_files():
    """The download buttons must not be a way to crash the page."""
    assert to_excel([])[:2] == b"PK"
    assert to_pdf([])[:5] == b"%PDF-"


@pytest.mark.parametrize("minor,currency,text", [
    (500_000, Currency.INR, "INR 5,000.00"),
    (1, Currency.INR, "INR 0.01"),
    (-150_000, Currency.INR, "INR -1,500.00"),
    (4_250, Currency.USD, "USD 42.50"),
    (100_000_000, Currency.USD, "USD 1,000,000.00"),
])
def test_pdf_money_never_needs_a_glyph_the_font_lacks(minor, currency, text):
    """fpdf's core fonts are latin-1, so a rupee sign would raise. The code
    says the same thing and always draws."""
    rendered = _money(minor, currency)
    assert rendered == text
    assert "₹" not in rendered
    rendered.encode("latin-1")  # raises if a core font could not draw it


def test_pdf_covers_both_currencies_without_mixing_them():
    pdf = to_pdf([
        entry(100_000, currency=Currency.INR),
        entry(20_000, currency=Currency.USD),
    ])
    assert pdf[:5] == b"%PDF-"
    assert len(pdf) > 1200


def test_rows_come_out_in_date_order():
    entries = [
        entry(100, date=date(2026, 5, 1)),
        entry(200, date=date(2026, 1, 1)),
        entry(300, date=date(2026, 3, 1)),
    ]
    assert [r[0] for r in _rows(entries)] == [
        date(2026, 1, 1), date(2026, 3, 1), date(2026, 5, 1)
    ]


# --------------------------------------------------- what the fonts can draw

class TestPdfSurvivesRealText:
    """The Download page crashed on the deployed app with
    `FPDFUnicodeEncodingException`. fpdf's core fonts are latin-1, and only
    `_money` was guarded — so a rupee sign in an *amount* was safe while the
    same sign in a *note* took the whole page down. Names, ledgers and notes
    are free text typed by a person or written by a model out of a document.
    """

    @pytest.mark.parametrize("text", [
        "gave ₹5,000 → for fees",       # the assistant writes both of these
        "📄 statement “Aug” — ok",       # emoji and curly quotes from a PDF
        "విహార్",                        # a name in Telugu
        "Ṡrī",                          # decomposable Latin
        "café • naïve",                 # latin-1 already, must survive intact
        "rent 1st–5th",                 # en dash
    ])
    def test_a_pdf_is_produced_whatever_the_text(self, text):
        from ledger.export import to_pdf as build

        pdf = build([entry(100_00, person=text, ledger=text, note=text)])
        assert pdf[:5] == b"%PDF-"

    @pytest.mark.parametrize("raw,expected", [
        ("₹5,000", "INR 5,000"),
        ("a → b", "a -> b"),
        ("“quoted”", '"quoted"'),
        ("it’s", "it's"),
        ("one—two", "one-two"),
        ("Ṡrī", "Sri"),
    ])
    def test_the_stand_ins_are_readable_not_question_marks(self, raw, expected):
        from ledger.export import _latin1

        assert _latin1(raw) == expected

    def test_latin1_text_is_left_exactly_alone(self):
        from ledger.export import _latin1

        assert _latin1("café naïve Zoë") == "café naïve Zoë"

    def test_everything_returned_can_actually_be_encoded(self):
        """The one property that matters: whatever comes out, fpdf can draw."""
        from ledger.export import _latin1

        for text in ["విహార్", "📄🎉", "日本語", "", None, "₹→“”"]:
            _latin1(text).encode("latin-1")

    def test_a_note_with_a_rupee_sign_no_longer_crashes(self):
        """The exact shape reported from the deployed app."""
        from ledger.export import to_pdf as build

        assert build([entry(500_00, note="paid ₹500 → cleared")])[:5] == b"%PDF-"
